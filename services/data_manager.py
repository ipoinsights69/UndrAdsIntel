import os
import json
import csv
import pandas as pd
import openpyxl
import re
from datetime import datetime
import copy


def get_all_apps():
    """
    Get all scraped app data
    """
    apps = []
    scraped_data_dir = os.getenv('SCRAPED_DATA_DIR', 'scraped_data')
    
    if os.path.exists(scraped_data_dir):
        for filename in os.listdir(scraped_data_dir):
            if filename.endswith('.json') and not filename.startswith('scrape_result_'):
                try:
                    with open(os.path.join(scraped_data_dir, filename), 'r') as f:
                        app_data = json.load(f)
                        if app_data.get('status') == 'success':
                            apps.append(app_data)
                except Exception as e:
                    print(f"Error loading {filename}: {str(e)}")
    
    return apps


def get_apps_by_developer(developer_name):
    """
    Get all apps by a specific developer
    """
    all_apps = get_all_apps()
    developer_name = developer_name.lower()
    result = []
    
    for app in all_apps:
        developer_info = app.get('developer_info') or {}
        dev_name = developer_info.get('name') or ''
        if dev_name.lower() == developer_name:
            result.append(app)
    
    return result


def search_apps(query):
    """
    Search apps by name, developer, or description
    """
    all_apps = get_all_apps()
    query = query.lower()
    
    results = []
    for app in all_apps:
        # Search in app name, developer name, app description, and package name
        developer_info = app.get('developer_info') or {}
        app_name = app.get('app_name') or ''
        dev_name = developer_info.get('name') or ''
        app_desc = app.get('app_description') or ''
        package = app.get('package_name') or ''
        
        if (query in app_name.lower() or
            query in dev_name.lower() or
            query in app_desc.lower() or
            query in package.lower()):
            results.append(app)
    
    return results


def export_data(format_type, filters=None, export_name=None):
    """Export data in the specified format with optional filters.
    
    Args:
        format_type (str): The format to export (csv, json, excel)
        filters (dict): Optional filters to apply to the data
        export_name (str, optional): Custom name for the export file
        
    Returns:
        str: The path to the exported file
    """
    # Create exports directory if it doesn't exist
    os.makedirs('exports', exist_ok=True)
    
    # Get all apps
    apps = get_all_apps()
    
    # Apply filters if provided
    if filters:
        filtered_apps = []
        for app in apps:
            # Initialize as True, will set to False if any filter doesn't match
            include_app = True
            
            # Ads filter
            if filters.get('ads_filter') and app.get('Ads'):
                if filters['ads_filter'] != app.get('Ads'):
                    include_app = False
            
            # Ranking filter
            if filters.get('ranking_filter') and app.get('Ranking'):
                if filters['ranking_filter'] != app.get('Ranking'):
                    include_app = False
                    
            
            # APK size filter
            if app.get('APK size'):
                apk_size_str = app.get('APK size', '').replace(' MB', '')
                try:
                    apk_size = float(apk_size_str)
                    
                    # Min APK size filter
                    if filters.get('apk_size_min') and float(filters['apk_size_min']) > apk_size:
                        include_app = False
                    
                    # Max APK size filter
                    if filters.get('apk_size_max') and float(filters['apk_size_max']) < apk_size:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the APK size, skip this filter
                    pass
            
            # Total downloads filter
            if app.get('Total downloads'):
                total_downloads_str = app.get('Total downloads', '')
                total_downloads_value = _parse_download_count(total_downloads_str)
                
                if filters.get('total_downloads_min') and filters.get('total_downloads_min_unit'):
                    min_value = float(filters['total_downloads_min'])
                    min_unit = filters['total_downloads_min_unit']
                    
                    # Convert to a common unit (thousands)
                    if min_unit == 'thousand':
                        min_value = min_value * 1000
                    elif min_unit == 'million':
                        min_value = min_value * 1000000
                    elif min_unit == 'billion':
                        min_value = min_value * 1000000000
                    
                    if total_downloads_value < min_value:
                        include_app = False
            
            # Recent downloads filter
            if app.get('Recent downloads'):
                recent_downloads_str = app.get('Recent downloads', '')
                recent_downloads_value = _parse_download_count(recent_downloads_str)
                
                if filters.get('recent_downloads_min') and filters.get('recent_downloads_min_unit'):
                    min_value = float(filters['recent_downloads_min'])
                    min_unit = filters['recent_downloads_min_unit']
                    
                    # Convert to a common unit (thousands)
                    if min_unit == 'thousand':
                        min_value = min_value * 1000
                    elif min_unit == 'million':
                        min_value = min_value * 1000000
                    elif min_unit == 'billion':
                        min_value = min_value * 1000000000
                    
                    if recent_downloads_value < min_value:
                        include_app = False
            
            # Android version filter
            if filters.get('android_version_min') and app.get('Designed for Android'):
                android_version_str = app.get('Designed for Android', '').replace('+', '')
                try:
                    app_version = float(android_version_str)
                    min_version = float(filters['android_version_min'])
                    
                    if app_version < min_version:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the version, skip this filter
                    pass
            
            # Rating filter (new)
            if app.get('Rating'):
                rating_value = _parse_rating(app.get('Rating', ''))
                
                # Min rating filter
                if filters.get('rating_min') and float(filters['rating_min']) > rating_value:
                    include_app = False
                
                # Max rating filter
                if filters.get('rating_max') and float(filters['rating_max']) < rating_value:
                    include_app = False
            
            # Developer filter
            if filters.get('developer') and app.get('developer_info', {}).get('name'):
                developer_name = app.get('developer_info', {}).get('name', '').lower()
                filter_developer = filters['developer'].lower()
                
                if filter_developer not in developer_name:
                    include_app = False
            
            # Country filter
            if filters.get('country_filter') and filters['country_filter']:
                app_country = app.get('country', '')
                if app_country != filters['country_filter']:
                    include_app = False
            
            # Category filter
            if filters.get('category_filter') and filters['category_filter']:
                app_category = app.get('category', '')
                if app_category != filters['category_filter']:
                    include_app = False
            
            # Is Ad filter
            if filters.get('is_ad_filter') and filters['is_ad_filter']:
                app_is_ad = app.get('is_ad', '')
                if app_is_ad != filters['is_ad_filter']:
                    include_app = False
            
            # Is Publisher filter
            if filters.get('is_pub_filter') and filters['is_pub_filter']:
                app_is_pub = app.get('is_pub', '')
                if app_is_pub != filters['is_pub_filter']:
                    include_app = False
            
            # Release Date filter
            if filters.get('release_date_min') and app.get('release_date'):
                try:
                    app_release_date = datetime.strptime(app.get('release_date', ''), '%Y-%m-%d')
                    min_date = datetime.strptime(filters['release_date_min'], '%Y-%m-%d')
                    if app_release_date < min_date:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the date, skip this filter
                    pass
                    
            if filters.get('release_date_max') and app.get('release_date'):
                try:
                    app_release_date = datetime.strptime(app.get('release_date', ''), '%Y-%m-%d')
                    max_date = datetime.strptime(filters['release_date_max'], '%Y-%m-%d')
                    if app_release_date > max_date:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the date, skip this filter
                    pass
            
            # Revenue USD filter
            if app.get('revenue_usd'):
                try:
                    revenue = float(app.get('revenue_usd', 0))
                    
                    # Min Revenue filter
                    if filters.get('revenue_usd_min') and float(filters['revenue_usd_min']) > revenue:
                        include_app = False
                    
                    # Max Revenue filter
                    if filters.get('revenue_usd_max') and float(filters['revenue_usd_max']) < revenue:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the revenue, skip this filter
                    pass
            
            # RPD (Revenue Per Download) filter
            if app.get('rpd'):
                try:
                    rpd = float(app.get('rpd', 0))
                    
                    # Min RPD filter
                    if filters.get('rpd_min') and float(filters['rpd_min']) > rpd:
                        include_app = False
                    
                    # Max RPD filter
                    if filters.get('rpd_max') and float(filters['rpd_max']) < rpd:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the RPD, skip this filter
                    pass
            
            # Downloads (numeric) filter
            if app.get('downloads'):
                try:
                    downloads = int(app.get('downloads', 0))
                    
                    # Min Downloads filter
                    if filters.get('downloads_min') and int(filters['downloads_min']) > downloads:
                        include_app = False
                    
                    # Max Downloads filter
                    if filters.get('downloads_max') and int(filters['downloads_max']) < downloads:
                        include_app = False
                except (ValueError, TypeError):
                    # If we can't parse the downloads, skip this filter
                    pass
                    
            # Ad Network Libraries filter
            if filters.get('ad_network_libraries') and app.get('api_details', {}).get('technologies', {}).get('Ad network libraries'):
                ad_libraries = app.get('api_details', {}).get('technologies', {}).get('Ad network libraries', [])
                if filters['ad_network_libraries'] not in ad_libraries:
                    include_app = False
            
            # Social Libraries filter
            if filters.get('social_libraries') and app.get('api_details', {}).get('technologies', {}).get('Social libraries'):
                social_libraries = app.get('api_details', {}).get('technologies', {}).get('Social libraries', [])
                if filters['social_libraries'] not in social_libraries:
                    include_app = False
            
            # Development Tools filter
            if filters.get('development_tools') and app.get('api_details', {}).get('technologies', {}).get('Development tools'):
                dev_tools = app.get('api_details', {}).get('technologies', {}).get('Development tools', [])
                if filters['development_tools'] not in dev_tools:
                    include_app = False
                    
            
            # If all filters passed, include the app
            if include_app:
                filtered_apps.append(app)
        
        # Replace the original apps list with the filtered list
        apps = filtered_apps
    
    # Process export options
    include_all_technologies = filters.get('include_all_technologies') == 'on' if filters else True
    include_latest_changelog = filters.get('include_latest_changelog') == 'on' if filters else True
    
    # Generate timestamp for filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format_type == 'csv':
        # Flatten nested data for CSV
        flattened_data = []
        for app in apps:
            flat_app = {}
            
            # Copy top-level fields
            for key, value in app.items():
                if key not in ['developer_info', 'api_details', 'changelog']:
                    flat_app[key] = value
            
            # Add developer info fields
            if 'developer_info' in app and app['developer_info']:
                for key, value in app['developer_info'].items():
                    flat_app[f'developer_{key}'] = value
            
            # Add technologies if requested
            if include_all_technologies and 'api_details' in app and app['api_details']:
                api_details = app['api_details']
                if 'technologies' in api_details:
                    technologies = api_details['technologies']
                    
                    # Add all technology categories
                    for category, tech_list in technologies.items():
                        if tech_list:
                            flat_app[f'tech_{category}'] = ', '.join(tech_list)
            
            # Add latest changelog if requested
            if include_latest_changelog and 'changelog' in app and app['changelog']:
                if len(app['changelog']) > 0:
                    latest = app['changelog'][0]  # Get the first (latest) changelog entry
                    flat_app['changelog_date'] = latest.get('date', '')
                    flat_app['changelog_type'] = latest.get('type', '')
                    flat_app['changelog_description'] = latest.get('description', '')
            
            flattened_data.append(flat_app)
        
        # Create CSV file
        if export_name:
            # Sanitize the export name to ensure it's a valid filename
            safe_name = re.sub(r'[^\w\-\.]', '_', export_name)
            filename = f'{safe_name}.csv'
        else:
            filename = f'app_data_{timestamp}.csv'
        filepath = os.path.join('exports', filename)
        
        if flattened_data:
            # Get all unique keys as fieldnames
            fieldnames = set()
            for app in flattened_data:
                fieldnames.update(app.keys())
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=sorted(fieldnames))
                writer.writeheader()
                writer.writerows(flattened_data)
        else:
            # Create empty CSV with default headers
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['No data found matching the filters'])
        
        return filename
    
    elif format_type == 'json':
        # Process apps for JSON export
        processed_apps = []
        
        for app in apps:
            processed_app = copy.deepcopy(app)
            
            # Handle technologies
            if not include_all_technologies and 'api_details' in processed_app:
                if 'technologies' in processed_app['api_details']:
                    processed_app['api_details'].pop('technologies', None)
            
            # Handle changelog
            if include_latest_changelog and 'changelog' in processed_app and processed_app['changelog']:
                if len(processed_app['changelog']) > 0:
                    processed_app['changelog'] = [processed_app['changelog'][0]]  # Keep only the latest entry
            
            processed_apps.append(processed_app)
        
        # Create JSON file
        if export_name:
            # Sanitize the export name to ensure it's a valid filename
            safe_name = re.sub(r'[^\w\-\.]', '_', export_name)
            filename = f'{safe_name}.json'
        else:
            filename = f'app_data_{timestamp}.json'
        filepath = os.path.join('exports', filename)
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(processed_apps, jsonfile, indent=2, ensure_ascii=False)
        
        return filename
    
    elif format_type == 'excel':
        # Create Excel file
        if export_name:
            # Sanitize the export name to ensure it's a valid filename
            safe_name = re.sub(r'[^\w\-\.]', '_', export_name)
            filename = f'{safe_name}.xlsx'
        else:
            filename = f'app_data_{timestamp}.xlsx'
        filepath = os.path.join('exports', filename)
        
        workbook = openpyxl.Workbook()
        
        # Main sheet for app data
        main_sheet = workbook.active
        main_sheet.title = 'Apps'
        
        # Flatten data for main sheet
        flattened_data = []
        for app in apps:
            flat_app = {}
            
            # Copy top-level fields
            for key, value in app.items():
                if key not in ['developer_info', 'api_details', 'changelog']:
                    flat_app[key] = value
            
            # Add developer info fields
            if 'developer_info' in app and app['developer_info']:
                for key, value in app['developer_info'].items():
                    flat_app[f'developer_{key}'] = value
            
            flattened_data.append(flat_app)
        
        if flattened_data:
            # Get all unique keys as fieldnames
            fieldnames = set()
            for app in flattened_data:
                fieldnames.update(app.keys())
            
            # Write header row
            header_row = sorted(fieldnames)
            main_sheet.append(header_row)
            
            # Write data rows
            for app in flattened_data:
                row = [app.get(field, '') for field in header_row]
                main_sheet.append(row)
            
            # Auto-adjust column widths
            for col in main_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                main_sheet.column_dimensions[column].width = adjusted_width
        else:
            # Create empty sheet with message
            main_sheet.append(['No data found matching the filters'])
        
        # Create additional sheets if needed
        if include_all_technologies:
            # Technologies sheet
            tech_sheet = workbook.create_sheet(title='Technologies')
            tech_sheet.append(['App Name', 'Package Name', 'Category', 'Technology'])
            
            row_index = 2
            for app in apps:
                if 'api_details' in app and app['api_details'] and 'technologies' in app['api_details']:
                    technologies = app['api_details']['technologies']
                    app_name = app.get('app_name', '')
                    package_name = app.get('package_name', '')
                    
                    for category, tech_list in technologies.items():
                        for tech in tech_list:
                            tech_sheet.append([app_name, package_name, category, tech])
                            row_index += 1
            
            # Auto-adjust column widths
            for col in tech_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                tech_sheet.column_dimensions[column].width = adjusted_width
        
        if include_latest_changelog:
            # Changelog sheet
            changelog_sheet = workbook.create_sheet(title='Changelog')
            changelog_sheet.append(['App Name', 'Package Name', 'Date', 'Type', 'Description'])
            
            row_index = 2
            for app in apps:
                if 'changelog' in app and app['changelog']:
                    app_name = app.get('app_name', '')
                    package_name = app.get('package_name', '')
                    
                    # Only include the latest changelog entry
                    if len(app['changelog']) > 0:
                        latest = app['changelog'][0]  # Get the first (latest) changelog entry
                        changelog_sheet.append([
                            app_name,
                            package_name,
                            latest.get('date', ''),
                            latest.get('type', ''),
                            latest.get('description', '')
                        ])
                        row_index += 1
            
            # Auto-adjust column widths
            for col in changelog_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                changelog_sheet.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        workbook.save(filepath)
        
        return filename
    
    else:
        raise ValueError(f"Unsupported export format: {format_type}")


def _parse_download_count(download_str):
    """Parse a download count string into a numeric value.
    
    Args:
        download_str (str): Download count string like "2.5 billion" or "32 million"
        
    Returns:
        float: The numeric value in absolute terms
    """
    if not download_str:
        return 0
    
    download_str = download_str.lower()
    
    # Extract the numeric part
    match = re.search(r'([\d.]+)', download_str)
    if not match:
        return 0
    
    value = float(match.group(1))
    
    # Apply multiplier based on unit
    if 'billion' in download_str:
        return value * 1000000000
    elif 'million' in download_str:
        return value * 1000000
    elif 'thousand' in download_str:
        return value * 1000
    else:
        return value


def _parse_rating(rating_str):
    """Parse a rating string into a numeric value.
    
    Args:
        rating_str (str): Rating string like "4.36 based on 160 million ratings"
        
    Returns:
        float: The numeric rating value
    """
    if not rating_str:
        return 0
    
    # Extract the numeric part (rating value)
    match = re.search(r'(\d+\.\d+)', rating_str)
    if not match:
        return 0
    
    return float(match.group(1))


def get_filter_options():
    """Get unique values for each filter field from the data.
    
    Returns:
        dict: A dictionary containing unique values for each filter field
    """
    apps = get_all_apps()
    
    # Initialize dictionaries to store unique values
    filter_options = {
        'countries': set(),
        'categories': set(),
        'ads_options': set(),
        'is_ad_options': set(),
        'is_pub_options': set(),
        'android_versions': set(),
        'rankings': set(),
        'ad_network_libraries': set(),
        'social_libraries': set(),
        'development_tools': set()
    }
    
    # Extract unique values from each app
    for app in apps:
        # Country
        if app.get('country'):
            filter_options['countries'].add(app.get('country'))
        
        # Category
        if app.get('category'):
            filter_options['categories'].add(app.get('category'))
        
        # Ads
        if app.get('Ads'):
            filter_options['ads_options'].add(app.get('Ads'))
        
        # Is Ad
        if app.get('is_ad'):
            filter_options['is_ad_options'].add(app.get('is_ad'))
        
        # Is Publisher
        if app.get('is_pub'):
            filter_options['is_pub_options'].add(app.get('is_pub'))
        
        # Android Version
        if app.get('Designed for Android'):
            filter_options['android_versions'].add(app.get('Designed for Android'))
        
        # Ranking
        if app.get('Ranking'):
            filter_options['rankings'].add(app.get('Ranking'))
        
        # Technology filters
        if app.get('api_details', {}).get('technologies', {}):
            technologies = app.get('api_details', {}).get('technologies', {})
            
            # Ad Network Libraries
            if technologies.get('Ad network libraries'):
                for lib in technologies.get('Ad network libraries', []):
                    filter_options['ad_network_libraries'].add(lib)
            
            # Social Libraries
            if technologies.get('Social libraries'):
                for lib in technologies.get('Social libraries', []):
                    filter_options['social_libraries'].add(lib)
            
            # Development Tools
            if technologies.get('Development tools'):
                for tool in technologies.get('Development tools', []):
                    filter_options['development_tools'].add(tool)
    
    # Convert sets to sorted lists for template rendering
    return {
        'countries': sorted(list(filter_options['countries'])),
        'categories': sorted(list(filter_options['categories'])),
        'ads_options': sorted(list(filter_options['ads_options'])),
        'is_ad_options': sorted(list(filter_options['is_ad_options'])),
        'is_pub_options': sorted(list(filter_options['is_pub_options'])),
        'android_versions': sorted(list(filter_options['android_versions']), key=lambda x: float(x.replace('+', '')) if x.replace('+', '').replace('.', '').isdigit() else 0),
        'rankings': sorted(list(filter_options['rankings'])),
        'ad_network_libraries': sorted(list(filter_options['ad_network_libraries'])),
        'social_libraries': sorted(list(filter_options['social_libraries'])),
        'development_tools': sorted(list(filter_options['development_tools']))
    }